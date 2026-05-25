from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


COLUMN_ALIASES = {
    "city": ["city", "город"],
    "problem_id": ["problem_id", "id", "candidate_id", "problem_key"],
    "title": ["problem_title", "title", "name", "problem", "label", "problem_name"],
    "score": ["complexity_score", "score", "complexity", "complexity_index", "index"],
    "frequency": ["frequency", "appeals_count", "count", "n_appeals", "mentions", "support"],
    "domains": ["domains", "domain", "top_domain", "thematic_areas", "sphere", "direction"],
    "topic": ["topic", "top_topic", "category", "theme"],
    "cross_domain_score": ["cross_domain_score", "cross_domain", "sphere_intersection"],
    "territory_count": ["territory_count", "locations_count", "address_count", "geo_count"],
    "months_count": ["months_count", "month_count", "period_count"],
    "executors_count": ["executors_count", "stakeholder_count", "actors_count"],
    "risk_score": ["risk_score", "risk", "emergency_score"],
}


def read_table(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()

    if suffix == ".csv":
        for encoding in ["utf-8-sig", "utf-8", "cp1251"]:
            try:
                return pd.read_csv(path, sep=None, engine="python", encoding=encoding)
            except Exception:
                pass
        return pd.read_csv(path)

    if suffix in [".xlsx", ".xls"]:
        return pd.read_excel(path)

    if suffix == ".jsonl":
        rows = []
        with path.open("r", encoding="utf-8") as f:
            for line in f:
                if line.strip():
                    rows.append(json.loads(line))
        return pd.DataFrame(rows)

    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, list):
            return pd.DataFrame(data)
        if isinstance(data, dict):
            for key in ["candidates", "problems", "items", "data"]:
                if key in data and isinstance(data[key], list):
                    return pd.DataFrame(data[key])
        return pd.DataFrame(data)

    raise ValueError(f"Неизвестный формат файла: {path}")


def find_column(df: pd.DataFrame, logical_name: str) -> Optional[str]:
    aliases = COLUMN_ALIASES.get(logical_name, [])
    normalized = {str(col).strip().lower(): col for col in df.columns}

    for alias in aliases:
        if alias.lower() in normalized:
            return normalized[alias.lower()]

    for col in df.columns:
        col_norm = str(col).strip().lower()
        for alias in aliases:
            if alias.lower() in col_norm:
                return col

    return None


def prepare_candidates(df: pd.DataFrame, city: str) -> pd.DataFrame:
    df = df.copy()

    city_col = find_column(df, "city")
    if city_col is None:
        df["city"] = city
    else:
        df["city"] = df[city_col].fillna(city)

    title_col = find_column(df, "title")
    score_col = find_column(df, "score")
    freq_col = find_column(df, "frequency")
    domains_col = find_column(df, "domains")
    topic_col = find_column(df, "topic")

    if title_col is None:
        df["problem_title"] = df.index.map(lambda i: f"Проблема {i + 1}")
    else:
        df["problem_title"] = df[title_col].astype(str)

    if score_col is None:
        df["complexity_score"] = 0.0
    else:
        df["complexity_score"] = pd.to_numeric(df[score_col], errors="coerce").fillna(0.0)

    if freq_col is None:
        df["frequency"] = 1
    else:
        df["frequency"] = pd.to_numeric(df[freq_col], errors="coerce").fillna(0).astype(int)

    if domains_col is None:
        df["domains"] = "Не указано"
    else:
        df["domains"] = df[domains_col].fillna("Не указано").astype(str)

    if topic_col is None:
        df["topic"] = df["domains"]
    else:
        df["topic"] = df[topic_col].fillna("Не указано").astype(str)

    for logical in ["cross_domain_score", "territory_count", "months_count", "executors_count", "risk_score"]:
        col = find_column(df, logical)
        if col is None:
            df[logical] = 0
        else:
            df[logical] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df["complexity_level"] = pd.cut(
        df["complexity_score"],
        bins=[-0.01, 0.4, 0.7, 1.01],
        labels=["Низкая", "Средняя", "Высокая"],
    )

    df["is_cross_domain"] = df["domains"].str.contains(";|,|\\+", regex=True) | (df["cross_domain_score"] > 0)

    return df


def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    rows = []

    rows.append(["Кандидатов всего", len(df)])
    rows.append(["Городов в отчете", df["city"].nunique()])
    rows.append(["Высокая комплексность", int((df["complexity_level"] == "Высокая").sum())])
    rows.append(["Средняя комплексность", int((df["complexity_level"] == "Средняя").sum())])
    rows.append(["Низкая комплексность", int((df["complexity_level"] == "Низкая").sum())])
    rows.append(["Межсферных проблем", int(df["is_cross_domain"].sum())])
    rows.append(["Средний индекс комплексности", round(float(df["complexity_score"].mean()), 3)])
    rows.append(["Медианный индекс комплексности", round(float(df["complexity_score"].median()), 3)])
    rows.append(["Суммарно подтверждающих обращений", int(df["frequency"].sum())])

    return pd.DataFrame(rows, columns=["Показатель", "Значение"])


def autosize_sheet(ws):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def style_workbook(path: Path):
    from openpyxl import load_workbook

    wb = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.freeze_panes = "A2"
        autosize_sheet(ws)

    if "Сравнение городов" in wb.sheetnames:
        ws = wb["Сравнение городов"]

        if ws.max_row >= 2:
            chart = BarChart()
            chart.title = "Кандидаты проблем по городам"
            chart.y_axis.title = "Количество"
            chart.x_axis.title = "Город"

            data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 8
            chart.width = 16

            ws.add_chart(chart, "H2")

    wb.save(path)


def build_markdown(df: pd.DataFrame, out_path: Path):
    total = len(df)
    high = int((df["complexity_level"] == "Высокая").sum())
    cross = int(df["is_cross_domain"].sum())
    avg_score = round(float(df["complexity_score"].mean()), 3)

    city_stats = (
        df.groupby("city")
        .agg(
            candidates=("problem_title", "count"),
            avg_score=("complexity_score", "mean"),
            appeals=("frequency", "sum"),
            high_complexity=("complexity_level", lambda s: int((s == "Высокая").sum())),
            cross_domain=("is_cross_domain", "sum"),
        )
        .reset_index()
    )

    top = df.sort_values(["complexity_score", "frequency"], ascending=False).head(10)

    lines = []
    lines.append("# Статистика по кандидатам комплексных городских проблем\n")
    lines.append(f"Всего выявлено **{total}** кандидатов комплексных городских проблем.")
    lines.append(f"Из них **{high}** имеют высокий индекс комплексности, **{cross}** являются межсферными.")
    lines.append(f"Средний индекс комплексности составил **{avg_score}**.\n")

    lines.append("## Сравнение городов\n")
    lines.append("| Город | Кандидатов | Средний индекс | Подтверждающих обращений | Высокая комплексность | Межсферные |")
    lines.append("|---|---:|---:|---:|---:|---:|")
    for _, row in city_stats.iterrows():
        lines.append(
            f"| {row['city']} | {int(row['candidates'])} | {row['avg_score']:.3f} | "
            f"{int(row['appeals'])} | {int(row['high_complexity'])} | {int(row['cross_domain'])} |"
        )

    lines.append("\n## Топ-10 проблем по комплексности\n")
    lines.append("| № | Город | Проблема | Индекс | Обращений | Сфера/категория |")
    lines.append("|---:|---|---|---:|---:|---|")
    for i, (_, row) in enumerate(top.iterrows(), start=1):
        title = str(row["problem_title"]).replace("|", " ")
        topic = str(row["topic"]).replace("|", " ")
        lines.append(
            f"| {i} | {row['city']} | {title} | {row['complexity_score']:.3f} | "
            f"{int(row['frequency'])} | {topic} |"
        )

    lines.append("\n## Интерпретация\n")
    lines.append(
        "Полученная статистика показывает не только количество выявленных проблем, "
        "но и их распределение по уровню комплексности, межсферности, частотности, "
        "территориальному охвату и устойчивости во времени."
    )

    out_path.write_text("\n".join(lines), encoding="utf-8")


def build_report(spb_path: Optional[Path], voronezh_path: Optional[Path], out_dir: Path):
    frames = []

    if spb_path and spb_path.exists():
        frames.append(prepare_candidates(read_table(spb_path), "Санкт-Петербург"))

    if voronezh_path and voronezh_path.exists():
        frames.append(prepare_candidates(read_table(voronezh_path), "Воронеж"))

    if not frames:
        raise FileNotFoundError("Не найден ни один файл кандидатов.")

    df = pd.concat(frames, ignore_index=True)

    out_dir.mkdir(parents=True, exist_ok=True)

    summary = build_summary(df)

    city_comparison = (
        df.groupby("city")
        .agg(
            candidates=("problem_title", "count"),
            avg_score=("complexity_score", "mean"),
            median_score=("complexity_score", "median"),
            appeals=("frequency", "sum"),
            high_complexity=("complexity_level", lambda s: int((s == "Высокая").sum())),
            cross_domain=("is_cross_domain", "sum"),
        )
        .reset_index()
    )

    level_distribution = (
        df.groupby(["city", "complexity_level"], observed=False)
        .size()
        .reset_index(name="count")
    )

    category_distribution = (
        df.groupby(["city", "topic"])
        .agg(candidates=("problem_title", "count"), appeals=("frequency", "sum"), avg_score=("complexity_score", "mean"))
        .reset_index()
        .sort_values(["city", "candidates", "appeals"], ascending=[True, False, False])
    )

    top_score = df.sort_values(["complexity_score", "frequency"], ascending=False).head(30)
    top_frequency = df.sort_values(["frequency", "complexity_score"], ascending=False).head(30)
    cross_domain = df[df["is_cross_domain"]].sort_values(["complexity_score", "frequency"], ascending=False)
    risks = df.sort_values(["risk_score", "complexity_score"], ascending=False).head(30)

    xlsx_path = out_dir / "candidate_statistics.xlsx"
    json_path = out_dir / "candidate_statistics.json"
    md_path = out_dir / "candidate_statistics.md"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Сводка", index=False)
        city_comparison.to_excel(writer, sheet_name="Сравнение городов", index=False)
        level_distribution.to_excel(writer, sheet_name="Уровни комплексности", index=False)
        top_score.to_excel(writer, sheet_name="Топ по индексу", index=False)
        top_frequency.to_excel(writer, sheet_name="Топ по обращениям", index=False)
        category_distribution.to_excel(writer, sheet_name="Категории", index=False)
        cross_domain.to_excel(writer, sheet_name="Межсферные", index=False)
        risks.to_excel(writer, sheet_name="Риски", index=False)
        df.to_excel(writer, sheet_name="Все кандидаты", index=False)

    style_workbook(xlsx_path)

    json_path.write_text(
        json.dumps(
            {
                "summary": summary.to_dict(orient="records"),
                "city_comparison": city_comparison.to_dict(orient="records"),
                "top_score": top_score.head(20).to_dict(orient="records"),
            },
            ensure_ascii=False,
            indent=2,
            default=str,
        ),
        encoding="utf-8",
    )

    build_markdown(df, md_path)

    print(f"Готово: {xlsx_path}")
    print(f"Готово: {md_path}")
    print(f"Готово: {json_path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--spb", type=Path, default=None)
    parser.add_argument("--voronezh", type=Path, default=None)
    parser.add_argument("--out-dir", type=Path, default=Path("outputs/candidate_report"))
    args = parser.parse_args()

    build_report(args.spb, args.voronezh, args.out_dir)


if __name__ == "__main__":
    main()